from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def test_demo_custom_publish_uses_shared_version_and_payload_helpers():
    modeling = Path("pricing_models/demo_custom_publish/modeling.py").read_text(encoding="utf-8")
    airflow_tasks = Path("pricing_models/demo_custom_publish/airflow_tasks.py").read_text(
        encoding="utf-8"
    )
    runner = Path("scripts/run_demo_custom_publish.py").read_text(encoding="utf-8")

    assert "completed_build_helpers import" in modeling
    assert "completed_model_build_payload(" in modeling
    assert "def existing_model_version_for_export" not in modeling
    assert "def next_trained_model_version" not in modeling
    assert "def effective_from_for_run" not in modeling
    assert "resolve_model_version_for_export" in airflow_tasks
    assert "resolve_model_version_for_export" in runner


def test_demo_effective_from_uses_run_date_not_hardcoded_string():
    from pricing_models.demo_custom_publish.data import training_table_for_run
    from pricing_pipeline.orchestration.run_context import (
        run_key_for_value,
    )

    assert run_key_for_value("manual__2026-06-04T23:15:00+00:00").startswith(
        "manual__20260604t2315000000_"
    )
    assert training_table_for_run("manual__2026-06-04T23:15:00+00:00").startswith(
        "DEMO_CUSTOM_PUBLISH_TRAINING_manual__20260604t2315000000_"
    )
    assert len(training_table_for_run("x" * 200)) <= 128


def test_demo_run_scoped_keys_are_collision_safe():
    from pricing_models.demo_custom_publish.data import (
        training_table_for_run,
    )
    from pricing_pipeline.orchestration.run_context import run_key_for_value

    punctuation_keys = {
        run_key_for_value("scheduled__2026-06-04T12:00:00+00:00"),
        run_key_for_value("scheduled__20260604T1200000000"),
    }
    long_prefix_keys = {
        run_key_for_value(("x" * 200) + "a"),
        run_key_for_value(("x" * 200) + "b"),
    }
    punctuation_tables = {
        training_table_for_run("scheduled__2026-06-04T12:00:00+00:00"),
        training_table_for_run("scheduled__20260604T1200000000"),
    }
    long_prefix_tables = {
        training_table_for_run(("x" * 200) + "a"),
        training_table_for_run(("x" * 200) + "b"),
    }

    assert len(punctuation_keys) == 2
    assert len(long_prefix_keys) == 2
    assert len(punctuation_tables) == 2
    assert len(long_prefix_tables) == 2
    assert all(len(key) <= 99 for key in punctuation_keys | long_prefix_keys)
    assert all(len(table) <= 128 for table in punctuation_tables | long_prefix_tables)


def test_demo_data_constants_describe_frame_manifest():
    from pricing_models.demo_custom_publish.data import (
        DATASET_NAME,
        PK_COLUMNS,
        SOURCE_SYSTEM,
        WEIGHT_COLUMN,
    )

    assert DATASET_NAME == "demo_custom_frequency_model_frame"
    assert SOURCE_SYSTEM == "demo_sql_server_staging"
    assert PK_COLUMNS == ("policy_id",)
    assert WEIGHT_COLUMN == "exposure"


def test_demo_training_export_returns_completed_build_payload(tmp_path: Path):
    from pricing_models.demo_custom_publish.data import (
        build_demo_training_frame,
    )
    from pricing_models.demo_custom_publish.modeling import (
        export_superglm_completed_build,
    )

    frame = build_demo_training_frame(row_count=180, seed=20260604)

    completed = export_superglm_completed_build(
        frame,
        output_dir=tmp_path,
        model_version="v1",
        effective_from="2026-06-04",
        created_by="pytest",
        export_id="demo-custom-publish-test",
        manifest_id="manifest-1",
        split_set_id="split-1",
    )

    workbook_path = Path(completed["rating_workbook_path"])
    model_path = Path(completed["model_artifact_path"])
    assert workbook_path.parent == tmp_path / "demo-custom-publish-test"
    summary_path = workbook_path.parent / "model_summary.txt"

    assert completed["model_version"] == "v1"
    assert completed["effective_from"] == "2026-06-04"
    assert completed["created_by"] == "pytest"
    assert completed["export_id"] == "demo-custom-publish-test"
    assert completed["manifest_id"] == "manifest-1"
    assert completed["split_set_id"] == "split-1"
    assert completed["mlflow_run_id"] is None
    assert completed["metrics"]["row_count"] == 180
    assert completed["metrics"]["deviance"] > 0
    assert workbook_path.exists()
    assert model_path.exists()
    assert summary_path.exists()

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    assert workbook.sheetnames == [
        "Rating Tables",
        "Discretization Impact",
        "Model Summary",
    ]
    assert workbook["Rating Tables"]["A2"].value == "Base"
    assert "SuperGLM Results" in summary_path.read_text(encoding="utf-8")
