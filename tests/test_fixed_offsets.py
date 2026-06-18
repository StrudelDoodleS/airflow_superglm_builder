from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from pricing_pipeline.models.config import ModelBuildConfig
from pricing_pipeline.models.spec import ModelExportResult
from pricing_pipeline.publishing.fixed_offsets import build_fixed_offset_staging_frame
from pricing_pipeline.publishing.lifecycle import PublishResult
from pricing_pipeline.publishing.publisher import ModelPublisher


def _write_workbook(path: Path, *, reference: float = 12.0, transform: str = "LOG_RATIO") -> None:
    workbook = Workbook()
    workbook.active.title = "Rating Tables"
    worksheet = workbook.create_sheet("Fixed Offsets")
    worksheet.append(
        [
            "Term",
            "Term Type",
            "Source Feature",
            "Transform",
            "Reference Value",
            "Coefficient",
            "Sequence",
            "Link Expression",
            "Response Multiplier",
        ]
    )
    worksheet.append(
        [
            "policy_term",
            "FIXED_OFFSET",
            "term",
            transform,
            reference,
            1.0,
            1,
            "1 * log(term / 12)",
            "(term / 12) ^ 1",
        ]
    )
    workbook.save(path)


def test_build_fixed_offset_staging_frame_reads_log_ratio_metadata(tmp_path: Path):
    workbook_path = tmp_path / "rating_tables.xlsx"
    _write_workbook(workbook_path)

    frame = build_fixed_offset_staging_frame(workbook_path, export_id="export-1")

    assert frame.to_dict("records") == [
        {
            "export_id": "export-1",
            "term_name": "policy_term",
            "source_feature_name": "term",
            "transform_type": "LOG_RATIO",
            "reference_value": 12.0,
            "coefficient": 1.0,
            "sequence_no": 1,
        }
    ]


def test_build_fixed_offset_staging_frame_returns_empty_when_sheet_is_absent(tmp_path: Path):
    workbook_path = tmp_path / "rating_tables.xlsx"
    workbook = Workbook()
    workbook.active.title = "Rating Tables"
    workbook.save(workbook_path)

    frame = build_fixed_offset_staging_frame(workbook_path, export_id="export-1")

    assert frame.empty
    assert list(frame.columns) == [
        "export_id",
        "term_name",
        "source_feature_name",
        "transform_type",
        "reference_value",
        "coefficient",
        "sequence_no",
    ]


def test_build_fixed_offset_staging_frame_rejects_unsupported_transform(tmp_path: Path):
    workbook_path = tmp_path / "rating_tables.xlsx"
    _write_workbook(workbook_path, transform="ARBITRARY_SQL")

    with pytest.raises(ValueError, match="Unsupported fixed offset transform"):
        build_fixed_offset_staging_frame(workbook_path, export_id="export-1")


def test_build_fixed_offset_staging_frame_rejects_non_positive_reference(tmp_path: Path):
    workbook_path = tmp_path / "rating_tables.xlsx"
    _write_workbook(workbook_path, reference=0.0)

    with pytest.raises(ValueError, match="strictly positive"):
        build_fixed_offset_staging_frame(workbook_path, export_id="export-1")


def test_fixed_offset_migration_scores_log_ratio_on_response_scale():
    migration = Path("db/migrations/V022__deployable_fixed_offsets.sql").read_text(
        encoding="utf-8"
    )

    assert "CREATE TABLE pricing_stg.STG_FIXED_OFFSET" in migration
    assert "CREATE TABLE pricing.PRICING_FIXED_OFFSET" in migration
    assert "CREATE OR ALTER VIEW pricing.V_CURRENT_FIXED_OFFSET" in migration
    assert "TR_PRICING_RATE_PACKAGE_COPY_FIXED_OFFSETS" in migration
    assert "fixed.coefficient * LOG" in migration
    assert "POWER(" in migration
    assert "THROW 50004" in migration
    assert "COUNT(*)\n            FROM pricing.V_CURRENT_FIXED_OFFSET" in migration


def test_model_publisher_stages_fixed_offsets_before_publish(tmp_path: Path, monkeypatch):
    calls = []
    workbook_path = tmp_path / "rating_tables.xlsx"
    workbook_path.touch()
    config = ModelBuildConfig(
        model_name="MTPL_FREQ",
        model_label="Motor frequency",
        target_name="ClaimNb",
        model_type="superglm_poisson",
        deployment_slot="MTPL_FREQ_UAT",
        default_package_status="PUBLISHED",
    )
    export = ModelExportResult(
        model_id=17,
        model_name="MTPL_FREQ",
        model_version="20260618",
        model_type="superglm_poisson",
        target_name="ClaimNb",
        deployment_slot="MTPL_FREQ_UAT",
        manifest_id="manifest-1",
        dag_id="dag",
        airflow_run_id="scheduled__2026-06-18",
        mlflow_run_id="mlflow-1",
        split_set_id=None,
        export_id="export-1",
        rating_workbook_path=str(workbook_path),
        effective_from="2026-06-18",
        created_by="airflow",
    )

    monkeypatch.setattr(
        "pricing_pipeline.publishing.publisher.validate_model_on_engine",
        lambda engine, config: calls.append("validate") or 17,
    )
    monkeypatch.setattr(
        "pricing_pipeline.publishing.publisher.stage_rating_export",
        lambda *args, **kwargs: calls.append("rating"),
    )
    monkeypatch.setattr(
        "pricing_pipeline.publishing.publisher.stage_fixed_offsets",
        lambda *args, **kwargs: calls.append("offset"),
    )
    monkeypatch.setattr(
        "pricing_pipeline.publishing.publisher.publish_rating_package",
        lambda *args, **kwargs: calls.append("publish")
        or PublishResult(
            mlflow_run_id="",
            export_id="export-1",
            rate_package_id=42,
            package_version=1,
            rating_workbook_path="",
        ),
    )

    result = ModelPublisher(object(), config).publish_training_export(export)

    assert result.rate_package_id == 42
    assert calls == ["validate", "rating", "offset", "publish"]
